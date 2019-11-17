import openpyxl
import re
wb = openpyxl.Workbook()
sheet = wb.active

def textToExcel(operating_file):
    c1 = sheet.cell(row=1, column=1)
    c1.value = "Time"
    c2 = sheet.cell(row=1, column=2)
    c2.value = "ucs1500"
    c3 = sheet.cell(row=1, column=3)
    c3.value = "ucs1501"
    c4 = sheet.cell(row=1, column=4)
    c4.value = "ucs1502"
    f = open("D:\python\cpu\\"+operating_file+".txt", "r")
    counter = 0
    row_ind = 2
    col_ind = 1
    for line in f:
        counter += 1
        if counter % 4 == 0:
            col_ind = 1
            date = re.findall(r'\d\d:\d\d:\d\d', line)
            i = sheet.cell(row=row_ind, column=col_ind)
            i.value = date[0]
            row_ind += 1
        else:
            col_ind += 1
            val = re.findall(r'(\d+)', line)
            i = sheet.cell(row=row_ind, column=col_ind)
            i.value = val[1]

    wb.save("D:\python\cpu\\"+operating_file + ".xlsx")
import os
arr_txt = [x for x in os.listdir('D:\python\cpu') if x.endswith(".txt")]
for file in arr_txt:
    filename = file.split(".")[0]
    textToExcel(filename)